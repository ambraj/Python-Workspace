import openpyxl

def convert_to_msec(time):
    time = time.strip()
    return (int(time[0]) * 60 + int(time[2:4])) * 1000 + int(time[-3:])

wb = openpyxl.load_workbook(filename = 'CatInVegas mobile sound Map 10-3-15.xlsx')
sheet = wb.get_sheet_by_name(name = 'Sheet1')

rr = 2


def read_sound_name(sound_name):
    sound_name = sound_name.lstrip()
    sound_name = sound_name.rstrip()
    sound_name = sound_name.replace(' ', '_')
    return sound_name

file = open("newfile.txt", "w")

while sheet.cell(row = rr, column = 2).value:

    string = sheet.cell(row = rr, column = 2).value.lower()
    sound_name = read_sound_name(string)

    string = sheet.cell(row = rr, column = 3).value
    start_time = convert_to_msec(string)

    string = sheet.cell(row = rr, column = 5).value
    duration = convert_to_msec(string)

    sound_slice = "\"" + sound_name + "\":\t\t\t\t\"frame\": [" + str(start_time) + ", " + str(duration) + "]},"

    print sound_slice

    file.write(sound_slice + "\n")

    rr += 1

file.close()

